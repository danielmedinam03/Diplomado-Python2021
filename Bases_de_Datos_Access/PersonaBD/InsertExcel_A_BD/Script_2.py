import os
import openpyxl
import pyodbc
from random import*

def conexionBD():
    # Ruta actual de la carpeta donde se esta trabajando
    # Conexion a la BD
    global conexion
    cadenaRutaBD = r'C:\Users\PC\Documents\Daniel\RETOVIBRA\Python\Bases_de_Datos_Access\PersonaBD\DataBase1.accdb;'
    driver = "Microsoft Access Driver (*.mdb, *.accdb)"
    cadena = (r'DRIVER={};DBQ={}').format(driver, cadenaRutaBD)
    conexion = pyodbc.connect(cadena)


conexionBD()

# Nombre del archivo txt para guardar los errores
nombreArchivo = "LogsError.txt"
archivo = open(nombreArchivo, "a")

# Documento de excel
ruta = os.getcwd()
documentoExcel = openpyxl.load_workbook(ruta + '\FacturaDetalles.xlsx')
sheets = documentoExcel.sheetnames
print(sheets)
hojaActiva = documentoExcel["FacturaDetalle"]
#print(hojaActiva.max_column)
#print(hojaActiva.max_row)

def menu():
    cursor = conexion.cursor()
    print("Insertar")
    n = 1
    ant = 0
    cont = 0
    for row in range(1, hojaActiva.max_row+1):
        valores = hojaActiva[f'A{row}:K{row}']
        clienteCC = hojaActiva[f'D{row}'].value
        """
            valores = hojaActiva[f'A{row}:K{row}']
            factura = hojaActiva[f'A{row}'].value
            fecha = hojaActiva[f'B{row}'].value
            # vendedor = hojaActiva[f'C{row}'].value
            clienteCC = hojaActiva[f'D{row}'].value
            codigoProduc = hojaActiva[f'E{row}'].value
            nombreProduc = hojaActiva[f'F{row}'].value
            cantidad = hojaActiva[f'G{row}'].value
            valorUnitario = hojaActiva[f'H{row}'].value
            # valorSuma = hojaActiva[f'I{row}'].value
            # iva = hojaActiva[f'J{row}'].value
            # netoPagar = hojaActiva[f'K{row}'].value
        """

        
    
        def Persona(clienteCC, cursor):
            cursor = conexion.cursor()

            selectPersona = "SELECT cNumeroDeIdentificacion FROM Persona"
            cursor.execute(selectPersona)
            rows = cursor.fetchall()
            #print(clienteCC)
            validacion = len(rows)
            #print(validacion)
            if validacion == 0:
                contPersona = 0
                nRandom = randrange(0,3,1)
                
                selectGenero = "SELECT cCodigoGenero FROM Genero WHERE cCodigoGenero = 'O'"
                cursor.execute(selectGenero)
                rowsGenero = cursor.fetchall()
                cursor.commit()
                print(rowsGenero)
                cursor.close()
                #genero = rowsGenero[nRandom]
                
                cursor = conexion.cursor()
                selectEstadoCvil = "SELECT cCodigoEstadoCivil FROM EstadoCivil WHERE =" + "S"
                cursor.execute(selectEstadoCvil)
                rowsEstadoCivil = cursor.fetchall()
                cursor.commit()
                print(rowsEstadoCivil)
                cursor.close()
                
                #estadoCivil = rowsEstadoCivil[nRandom]
                
                cursor = conexion.cursor()
                selectEstrato = "SELECT cCodigoEstrato FROM Estrato WHERE cCodigoEstrato = 'M'" 
                cursor.execute(selectEstrato)
                rowsEstrato = cursor.fetchall()
                cursor.commit()
                print(rowsEstrato)
                cursor.close()
                #estrato = rowsEstrato[nRandom]
                
                insertPersona = "INSERT INTO Persona(cNumeroDeIdentificacion, cNombre1, cNombre2, cApellido1, cApellido2, dFechaNacimiento, cCelular, cCodigoGenero, cTelefono, cCorreo, cDirreccion, cCodigoTipoID, cCodigoEstadoCivil, cCodigoEstrato) VALUES ("+str(clienteCC) +","+ "Daniel" + str(contPersona)+ ","+"Felipe"+ str(contPersona)+ ","+"Medina"+ str(contPersona)+ "," + "Moreno"+ str(contPersona)+ "," + "32221480-"+ str(contPersona)+","+ str(rowsGenero) + "," + "2625573-" + str(contPersona) +","+"Correo@gmail.com-" + str(contPersona) + ","+ "Direccion-" + str(contPersona) + ","+"CC" +","+str(rowsEstadoCivil) + ","+ str(rowsEstrato) + ")"
                cursor.execute(insertPersona)
                cursor.commit()
        Persona(clienteCC, cursor)

        # print(type(valores))
menu()
