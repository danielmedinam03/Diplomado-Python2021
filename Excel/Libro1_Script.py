import os
import openpyxl

lcRutaBase = os.getcwd()
print(lcRutaBase)

#lcRutaText = lcRutaBase + "\Text"
lnMaxFilas = 0
libro = 'Libro1_Excel_1.xlsx'

# Abre un libro de excel
excel_document = openpyxl.load_workbook(lcRutaBase + '\\' + libro)
#print (type(excel_document))

# Conocer los nombres de las hojas del libro
#print(excel_document.sheetnames)
"""
# Ingreso al contenido de una libro de excel
sheet = excel_document['Hoja1']
print(sheet.title)

lnMaxFilas = sheet.max_row
lnMaxCol = sheet.max_column

print (sheet['A2'].value)
print(sheet.cell(2,2).value)
"""

#excel_document.create_sheet("Danieljeje",1)

print(excel_document.sheetnames)

#Añade la hoja al documento
#hoja2 = excel_document.create_sheet("JEJEJEJE")


# Muestra los nombres de las hojas
print(excel_document.sheetnames)

#excel_document.save(hoja2)

#Guardar los cambios dentro del excel SE DEBE CERRAR EL EXCEL
excel_document.save(libro)

#Eliminar hojas en excel
excel_document.remove(excel_document["Daniel"])
excel_document.save(libro)

