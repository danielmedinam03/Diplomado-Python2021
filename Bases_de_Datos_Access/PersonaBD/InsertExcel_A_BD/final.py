import os
import openpyxl
import pyodbc
from random import*

    # Ruta actual de la carpeta donde se esta trabajando
    # Conexion a la BD
ruta = os.getcwd()

cadenaRutaBD = ruta + '\DataBase1.accdb;'
driver = "Microsoft Access Driver (*.mdb, *.accdb)"
cadena = (r'DRIVER={};DBQ={}').format(driver, cadenaRutaBD)

conexion = pyodbc.connect(cadena)

rutaTXT= ruta + '\logsErrores.txt'
errorRegistro = "--Registro ya existe--"

archivo = open(rutaTXT, "a")

documentoExcel = openpyxl.load_workbook(ruta + '\FacturaDetalles.xlsx')
sheets = documentoExcel.sheetnames
print(sheets)
hojaActiva = documentoExcel["FacturaDetalle"]

ant = 0   # variable que sirve para validar con respecto a la fila anterior
item = 1
for row in range(2, hojaActiva.max_row+1):

    # TABLA PERSONA
    # ---------------------------
    PKidPersona = hojaActiva['D'+str(row)].value
    
    selectIdPersona = "SELECT cNumeroDeIdentificacion FROM Persona WHERE cNumeroDeIdentificacion = '{}'".format(
        PKidPersona)

    cursor = conexion.cursor()
    cursor.execute(selectIdPersona)
    rowsIDpersona = cursor.fetchall()

    nombre1 = "Daniel - " + str(row)
    apellido1 = "Medina - " + str(row)
    celular = "3222148016 - " + str(row)
    correo = "daniel@gmail.com - " + str(row)
    direccion = "Cra " + str(row)
    tipoId = "CC"

    verificador = rowsIDpersona.__len__()

    if verificador == 0:
        insertPersonas = "INSERT INTO Persona(cNumeroDeIdentificacion, cNombre1, cApellido1, cCelular,cCorreo, cDirreccion, cCodigoTipoID) VALUES ('{}','{}','{}','{}','{}','{}','{}')".format(
            PKidPersona, nombre1, apellido1, celular, correo, direccion, tipoId)
        cursor.execute(insertPersonas)
        cursor.commit()
    else:
        archivo.write("\n" + errorRegistro + str(PKidPersona))

    # TABLA PRODUCTO
    # ------------------------------

    codigoProducto = hojaActiva['E'+str(row)].value
    selectCodProducto = "SELECT cCodProducto FROM Productos WHERE cCodProducto = '{}'".format(
        codigoProducto)

    cursor = conexion.cursor()
    cursor.execute(selectCodProducto)
    rowsCodigoProducto = cursor.fetchall()

    nombreProducto = hojaActiva['F'+str(row)].value
    verificador = rowsCodigoProducto.__len__()

    if verificador == 0:
        insertProducto = "INSERT INTO Productos(cCodProducto, cNombreProducto) VALUES ('{}','{}')".format(
            codigoProducto, nombreProducto)
        cursor = conexion.cursor()
        cursor.execute(insertProducto)
        cursor.commit()
    else:
        archivo.write("\n" + errorRegistro + str(codigoProducto))
        
    # TABLA MOVIMIENTO ENCABEZADO
    # ----------------------------------

    numMovimiento = hojaActiva[f'A{row}'].value

    IDcliente = hojaActiva[f'D{row}'].value

    selectCodMovimiento = "SELECT * FROM Movi_Encab WHERE nNum_Mov = '{}' and cNumeroIdentificacion = '{}'".format(
        numMovimiento, IDcliente)

    cursor = conexion.cursor()
    cursor.execute(selectCodMovimiento)
    rowsCodigoMovimiento = cursor.fetchall()

    verificador = rowsCodigoMovimiento.__len__()

    if verificador == 0:
        
        insertMovimiento = "INSERT INTO Movi_Encab(cCodMov, nNum_Mov, cNumeroIdentificacion) VALUES ('FACT','{}','{}')".format(
            numMovimiento, IDcliente)
        cursor = conexion.cursor()
        cursor.execute(insertMovimiento)
        cursor.commit()
    else:
        archivo.write("\n" + errorRegistro + str(numMovimiento))
 
    # TABLA MOVIMIENTO DETALLE
    # ----------------------------------
    codigoMovimiento = hojaActiva["A"+str(row)].value
    codigoProducto = hojaActiva['E'+str(row)].value
    valorUnitario = str(hojaActiva['H'+str(row)].value)
    cantidad = str(hojaActiva['G'+str(row)].value)
    
    selectDetalle = "SELECT * FROM Movi_Detalle"
    cursor = conexion.cursor()
    cursor.execute(selectDetalle)
    data=cursor.fetchall()
    
    sig = hojaActiva["A"+str(row)].value
    
    if ant == sig:
        item += 1
        insertDetalle = "INSERT INTO Movi_Detalle(cCodMov, nNum_Mov, nItem, cCodProducto, nValorUnitario,nCantidad) VALUES ('FACT','{}','{}','{}','{}','{}')".format(
                codigoMovimiento, item, codigoProducto, valorUnitario, cantidad)
        cursor = conexion.cursor()
        cursor.execute(insertDetalle)
        cursor.commit()
        
    else:
        
        ant = sig
        if item >0 :
            item = 1
        insertDetalle = "INSERT INTO Movi_Detalle(cCodMov, nNum_Mov, nItem, cCodProducto, nValorUnitario,nCantidad) VALUES ('FACT','{}','{}','{}','{}','{}')".format(
                codigoMovimiento, item, codigoProducto, valorUnitario, cantidad)
        cursor = conexion.cursor()
        cursor.execute(insertDetalle)
        cursor.commit()

    
cursor.close()   

conexion.close()
    
