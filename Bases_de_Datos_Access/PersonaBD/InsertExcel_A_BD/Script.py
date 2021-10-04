import os
import openpyxl
import pyodbc
#Try para manejar los errores y poder insertarlos en el txt

try:
    def conexionBD():
    #Ruta actual de la carpeta donde se esta trabajando 
    #Conexion a la BD
        global conexion
        cadenaRutaBD=r'C:\Users\PC\Documents\Daniel\RETOVIBRA\Python\Bases_de_Datos_Access\PersonaBD\DataBase1.accdb;'
        driver = "Microsoft Access Driver (*.mdb, *.accdb)" 
        cadena = (r'DRIVER={};DBQ={}').format(driver,cadenaRutaBD)
        conexion = pyodbc.connect(cadena)
    conexionBD()
    
    #Nombre del archivo txt para guardar los errores
    nombreArchivo="LogsError.txt"
    archivo=open(nombreArchivo,"a")

    #Documento de excel
    ruta = os.getcwd()
    documentoExcel = openpyxl.load_workbook(ruta + '\FacturaDetalles.xlsx')
    sheets = documentoExcel.sheetnames
    print(sheets)
    hojaActiva = documentoExcel["FacturaDetalle"]
    print(hojaActiva.max_column)
    print(hojaActiva.max_row)
    
    
    def menu(conexion):
        print("Insertar")
        n =1
        ant=0
        cont=0
        for row in range(1,hojaActiva.max_row+1):
            """
            valores = hojaActiva[f'A{row}:K{row}']
            factura = hojaActiva[f'A{row}'].value
            fecha = hojaActiva[f'B{row}'].value
            #vendedor = hojaActiva[f'C{row}'].value
            clienteCC = hojaActiva[f'D{row}'].value
            codigoProduc = hojaActiva[f'E{row}'].value
            nombreProduc = hojaActiva[f'F{row}'].value
            cantidad = hojaActiva[f'G{row}'].value
            valorUnitario = hojaActiva[f'H{row}'].value
            #valorSuma = hojaActiva[f'I{row}'].value
            #iva = hojaActiva[f'J{row}'].value
            #netoPagar = hojaActiva[f'K{row}'].value
            """
            cursor = conexion.cursor()
            
            valores = hojaActiva[f'A{row}:K{row}']
            clienteCC = hojaActiva[f'D{row}'].value
            def Persona(clienteCC,cursor):
                print(clienteCC)
                
                selectPersona = "SELECT cNumeroDeIdentificacion FROM Persona"
                cursor.execute(selectPersona)
                rows = cursor.fetchall()
                validacion = len(rows)
                if validacion == 0:
                    cursor.conexion()
                    selectGenero = "SELECT cCodigoGenero FROM Genero"
                    selectEstrato = "SELECT cCodigoEstrato FROM Estrato" 
                    selectEstadoCvil = "SELECT cCodigoEstadoCivil FROM EstadoCivil" 
                    
                    insertPersona = "INSERT INTO Persona(cNumeroDeIdentificacion, cNombre1, cNombre2, cApellido1, cApellido2, dFechaNacimiento, cCelular, cCodigoGenero, cTelefono, cCorreo, cDirreccion, cCodigoTipoID, cCodigoEstadoCivil, cCodigoEstrato) VALUES ()"
            Persona(clienteCC,cursor)

        #print(type(valores))    
            
        
    menu(conexion)
    
    
        
        
#Manejo de excepciones para poder manipular los errores  
except Exception as e:
    print(e)
    e = str(e)
    archivo.writelines(e + "\n" + "\r")
        

archivo.close()