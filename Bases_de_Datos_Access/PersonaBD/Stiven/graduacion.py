import os
import pyodbc
import openpyxl
import random

ruta = os.getcwd()
print(ruta)
#[x for x in pyodbc.drivers() if x.startswith("Microsoft Access Driver")]

cadena = (r'DRIVER={Microsoft Access Driver (*.mdb, *.accdb)};'r'DBQ=' +
          ruta + r'\database\Tablas_practica.accdb;')
conexion = pyodbc.connect(cadena)

rutafile = os.getcwd() + "\\txt\\"

ubicacionfile = rutafile + "log.txt"
errores = ("------------------------ El codigo YA EXISTE ------------------------")
archivo = open(rutafile + "log.txt", 'a')


excel_document = openpyxl.load_workbook(ruta + '\FacturaDetalles.xlsx')
#hoja = excel_document.create_sheet("Hoja")

# Ingreso al contenido de una libro de excel
sheet = excel_document['FacturaDetalle']
print(sheet)

numColum = sheet.max_column
numFila = sheet.max_row

print(sheet.cell(2, 2).value)

numFact = 0
item = 1


# for i in range(1600):

# for columna in range(1, numColum+1):


"""for fila in range(1, numFila+1):
        p = {(sheet.cell(fila, columna).value)}
        print(p)
"""


for fila in range(2, numFila+1):

    # ---------------------------------------- REGISTROS CLIENTES ------------------------------------------------
    cont = 1
    # cNombre1, cApellido1, dDireccion_persona, cCelular, cEmail
    idcliente = sheet["D"+str(fila)].value
    IcSelect = "SELECT * FROM Personas WHERE cNumero_identificacion = '{}'".format(
        idcliente)
    Tabla = conexion.cursor()
    Tabla.execute(IcSelect)
    datos = Tabla.fetchall()

    nombre = "Steven" + str(fila)
    apellido = "Gonzalez" + str(fila)
    direccion = "Casa" + str(fila)
    celular = random.randint(3000000000, 3999999999)
    email = "Email" + str(fila) + "@gmail.com"
    genero = ['Masculino', 'Femenino', 'Otro']
    sexo = random.choice(genero)

    verificador = datos.__len__()
    if verificador > 0:
        archivo.write("\n" + errores)

    else:
        IcInsert = "INSERT INTO Personas(cNumero_identificacion, cNombre1, cApellido1, dDireccion_persona, cCelular, cEmail) VALUES('{}', '{}', '{}', '{}', '{}', '{}')".format(
            idcliente, nombre, apellido, direccion, celular, email)
        Tabla = conexion.cursor()
        Tabla.execute(IcInsert)
        Tabla.commit()
        print(nombre)

    cont += 1

    # ------------------------------------- PRODUCTOS -----------------------------------------------

    idPro = sheet["E"+str(fila)].value
    desPro = sheet["F"+str(fila)].value
    precioPro = sheet["H"+str(fila)].value

    IcSelect = "SELECT * FROM Productos WHERE cCodigoProducto = '{}'".format(
        idPro)
    Tabla = conexion.cursor()
    Tabla.execute(IcSelect)
    datos = Tabla.fetchall()

    verificador = datos.__len__()
    if verificador > 0:
        archivo.write("\n" + errores)

    else:
        IcInsert = "INSERT INTO Productos(cCodigoProducto, descripcion, precio) VALUES('{}', '{}', '{}')".format(
            idPro, desPro, precioPro)
        Tabla = conexion.cursor()
        Tabla.execute(IcInsert)
        Tabla.commit()
        print(nombre)

    # ------------------------------------- MOVIMIENTO ENCABEZADO -----------------------------------------------

    codMovi = sheet["A"+str(fila)].value
    codCli = sheet["D"+str(fila)].value

    IcSelect = "SELECT * FROM Mov_encab WHERE  num_Mov = '{}' and cNumero_identificacion = '{}'".format(
        codMovi, codCli)
    Tabla = conexion.cursor()
    Tabla.execute(IcSelect)
    datos = Tabla.fetchall()

    verificador = datos.__len__()
    if verificador > 0:
        archivo.write("\n" + errores)
    else:
        IcInsert = "INSERT INTO Mov_encab(codMov, num_Mov, cNumero_identificacion) VALUES('FACTURA','{}', '{}')".format(
            codMovi, codCli)
        Tabla = conexion.cursor()
        Tabla.execute(IcInsert)
        Tabla.commit()
        print(nombre)

    # ------------------------------------- MOVIMIENTO DETALLE -----------------------------------------------

    codMovi = sheet["A"+str(fila)].value
    idPro = sheet["E"+str(fila)].value
    precioPro = sheet["H"+str(fila)].value
    cantPro = sheet["G"+str(fila)].value

    IcSelect = "SELECT * FROM Mov_detalle "
    Tabla = conexion.cursor()
    Tabla.execute(IcSelect)
    datos = Tabla.fetchall()

    siguiente = sheet["A"+str(fila)].value

    if numFact == siguiente:
        item = item + 1
        IcInsert = "INSERT INTO Mov_detalle(codMov, num_Mov, item, cCodigoProducto, valorUnitario, cantidad) VALUES('FACTURA','{}', '{}', '{}', '{}', '{}')".format(
            codMovi, item, idPro, precioPro, cantPro)
        Tabla = conexion.cursor()
        Tabla.execute(IcInsert)
        Tabla.commit()

        

    else:
        numFact = siguiente
        if item > 0:
                item = 1
        IcInsert = "INSERT INTO Mov_detalle(codMov, num_Mov, item, cCodigoProducto, valorUnitario, cantidad) VALUES('FACTURA','{}', '{}', '{}', '{}', '{}')".format(
            codMovi, item, idPro, precioPro, cantPro)
        Tabla = conexion.cursor()
        Tabla.execute(IcInsert)
        Tabla.commit()





"""
    verificador = datos.__len__()
    if verificador > 0:
        archivo.write("\n" + errores)
"""

archivo.close

Tabla.close()
conexion.close()
