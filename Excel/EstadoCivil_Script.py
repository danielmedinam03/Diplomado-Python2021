import os
import openpyxl
import pyodbc


lcRutaBase = os.getcwd()
#lcRutaText = lcRutaBase + "\Text"
lnMaxFilas = 0
#print(lcRutaBase)

#Cadena de ruta de la BD
cadenaRutaBD=r'C:\Users\PC\Documents\Daniel\RETOVIBRA\Python\Bases_de_Datos_Access\PersonaBD\DataBase1.accdb;'
driver = "Microsoft Access Driver (*.mdb, *.accdb)" 
cadena = (r'DRIVER={};DBQ={}').format(driver,cadenaRutaBD)
conexion = pyodbc.connect(cadena)


# Abre un libro de excel
excel_document = openpyxl.load_workbook(lcRutaBase + '\EstadoCivil.xlsx')
#Libro de excel
libro="EstadoCivil.xlsx"
#print (type(excel_document))

# Conocer los nombres de las hojas del libro
print(excel_document.sheetnames[3])
"""
# Ingreso al contenido de una libro de excel
sheet = excel_document['EstadoCivil']
print(sheet.title)

lnMaxFilas = sheet.max_row
lnMaxCol = sheet.max_column

#sheet['A1'].value("Hola")

print (sheet['A2'].value)
print(sheet.cell(2,2).value)
"""

#crearHojaNueva = excel_document.create_sheet("DanielEstadoCivil")

excel_document.save(libro)

cursor = conexion.cursor()

selectEstadoCivil = "SELECT cCodigoEstadoCivil, cDescripcionEstadoCivil FROM EstadoCivil"

cursor.execute(selectEstadoCivil)

rows = cursor.fetchall()
#print(rows)
lista = []
for i in rows:
    for j in i:
        #print(j)
        lista.append(j)
        #value= f"j"
        #excel_document[value]=value
        
excel_document.save(libro)
print(lista)
#hoja = excel_document.sheetnames[3]
hojaDanielEstadoCivil = excel_document.active
print(hojaDanielEstadoCivil)

hojaDanielEstadoCivil["A1"] = lista[0]
hojaDanielEstadoCivil["B1"] = lista[1]

#celdaA1 = hojaDanielEstadoCivil(row=1, column =1,value=lista[0])

excel_document.save(libro)
